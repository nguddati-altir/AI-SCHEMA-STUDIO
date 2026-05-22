from __future__ import annotations
import io
import json
import datetime
from typing import Any
from openpyxl import load_workbook
from .models import Schema, Column


def validate(schema: Schema, file_bytes: bytes) -> dict:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    table_lookup = {t.name: t for t in schema.tables}
    sheets_report: dict[str, dict] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if sheet_name not in table_lookup:
            sheets_report[sheet_name] = {
                "status": "skipped",
                "reason": f"No table named '{sheet_name}' in the schema.",
                "rows": max(0, len(rows) - 1),
            }
            continue

        table = table_lookup[sheet_name]
        col_names = [c.name for c in table.columns]

        if not rows:
            sheets_report[sheet_name] = {"status": "error", "reason": "Sheet is empty.", "rows": 0}
            continue

        headers = [str(h) if h is not None else "" for h in rows[0]]
        missing = [c for c in col_names if c not in headers]
        extra = [h for h in headers if h and h not in col_names]
        data_rows = rows[1:]

        # Preview (stringify values for JSON safety)
        preview = []
        for r in data_rows[:25]:
            preview.append([_to_jsonable(v) for v in r])

        report: dict[str, Any] = {
            "status": "ok" if not missing and not extra else "mismatch",
            "rows": len(data_rows),
            "headers": headers,
            "missing_columns": missing,
            "extra_columns": extra,
            "preview": preview,
        }
        sheets_report[sheet_name] = report

    # Tables defined in schema but not in workbook
    for tname in table_lookup:
        if tname not in sheets_report:
            sheets_report[tname] = {"status": "absent", "reason": "No sheet in workbook.", "rows": 0}

    overall_ok = all(r.get("status") in {"ok", "absent"} for r in sheets_report.values())
    return {"ok": overall_ok, "sheets": sheets_report}


def generate_inserts(schema: Schema, file_bytes: bytes, dialect: str = "postgres") -> str:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    table_lookup = {t.name: t for t in schema.tables}
    qi = lambda n: f'"{n}"' if dialect == "postgres" else f"`{n}`"
    out: list[str] = [f"-- Generated INSERT statements ({dialect})"]

    for sheet_name in wb.sheetnames:
        if sheet_name not in table_lookup:
            continue
        table = table_lookup[sheet_name]
        col_lookup = {c.name: c for c in table.columns}
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        headers = [str(h) if h is not None else "" for h in rows[0]]
        usable_headers = [h for h in headers if h in col_lookup]
        if not usable_headers:
            continue

        out.append(f"\n-- {sheet_name}")
        for r in rows[1:]:
            row_dict = dict(zip(headers, r))
            values = [_format_sql(row_dict.get(h), col_lookup[h]) for h in usable_headers]
            cols = ", ".join(qi(h) for h in usable_headers)
            vals = ", ".join(values)
            out.append(f"INSERT INTO {qi(sheet_name)} ({cols}) VALUES ({vals});")

    return "\n".join(out) + "\n"


def _to_jsonable(v: Any) -> Any:
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.isoformat()
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    return str(v)


def _format_sql(v: Any, col: Column) -> str:
    if v is None or v == "":
        return "NULL"
    if col.type in {"int", "bigint"}:
        try:
            return str(int(v))
        except (TypeError, ValueError):
            return "NULL"
    if col.type in {"float", "decimal"}:
        try:
            return str(float(v))
        except (TypeError, ValueError):
            return "NULL"
    if col.type == "boolean":
        truthy = str(v).strip().lower() in {"1", "true", "yes", "y", "t"}
        return "TRUE" if truthy else "FALSE"
    if col.type == "json":
        if isinstance(v, (dict, list)):
            payload = json.dumps(v)
        elif isinstance(v, str):
            payload = v
        else:
            payload = json.dumps(str(v))
        return "'" + payload.replace("'", "''") + "'"
    if col.type in {"date", "datetime"}:
        if isinstance(v, (datetime.datetime, datetime.date)):
            return "'" + v.isoformat(sep=" ") + "'"
        return "'" + str(v).replace("'", "''") + "'"
    return "'" + str(v).replace("'", "''") + "'"
